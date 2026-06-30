"""Compliance Mapping Tool - FastAPI Backend."""

from contextlib import asynccontextmanager
from io import BytesIO
from typing import Optional

from fastapi import FastAPI, Depends, UploadFile, File, Form, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, or_, and_, case
from sqlalchemy.ext.asyncio import AsyncSession

from database import (
    init_db, get_session,
    Framework, Control, Mapping, VersionChange,
)
from document_parser import parse_uploaded_bytes, list_parsers


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class FrameworkOut(BaseModel):
    id: int
    name: str
    short_name: str
    version: str
    description: str
    is_active: bool
    control_count: int = 0

class ControlOut(BaseModel):
    id: int
    framework_id: int
    control_id: str
    title: str
    description: str
    category: str
    framework_short_name: str = ""

class MappingOut(BaseModel):
    id: int
    control_id: str
    title: str
    description: str
    category: str
    framework_short_name: str
    framework_id: int
    confidence: float
    source_type: str
    source_document: str
    notes: str = ""
    implementation_status: str = "not_assessed"
    owner: str = ""
    review_date: str = ""
    evidence_notes: str = ""

class MappingDetail(BaseModel):
    source: ControlOut
    mappings: list[MappingOut]


class MappingCreate(BaseModel):
    source_control_id: int
    target_control_id: int
    confidence: float = 1.0
    source_type: str = "manual"
    source_document: str = ""
    notes: str = ""


class MappingUpdate(BaseModel):
    confidence: Optional[float] = None
    source_type: Optional[str] = None
    notes: Optional[str] = None
    source_document: Optional[str] = None
    implementation_status: Optional[str] = None
    owner: Optional[str] = None
    review_date: Optional[str] = None
    evidence_notes: Optional[str] = None


class ControlSearchOut(BaseModel):
    total: int
    limit: int
    offset: int
    items: list[ControlOut]

class CoverageOut(BaseModel):
    source_framework: str
    target_framework: str
    total_source_controls: int
    mapped_controls: int
    unmapped_controls: int
    coverage_percentage: float
    unmapped_control_ids: list[dict]
    gap_controls: list[dict]

class VersionChangeOut(BaseModel):
    id: int
    old_version: str
    new_version: str
    change_type: str
    old_control_id: str
    new_control_id: str
    description: str
    category: str

class ParseResult(BaseModel):
    success: bool
    controls: list[dict] = []
    mappings: list[dict] = []
    raw_text: str = ""
    error: str = ""

class ImportRequest(BaseModel):
    doc_type: str = ""
    source_framework_id: int = 0
    target_framework_id: int = 0
    document_year: str = ""
    source_document: str = ""
    controls: list[dict] = []
    mappings: list[dict] = []

class ImportResult(BaseModel):
    success: bool
    controls_added: int = 0
    mappings_added: int = 0
    error: str = ""


# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

@asynccontextmanager
async def lifespan(application: FastAPI):
    await init_db()
    yield

app = FastAPI(
    title="Compliance Mapping Tool",
    description="Map controls between ISO 27001, BSI IT-Grundschutz, and C5",
    version="1.0.0",
    lifespan=lifespan,
)


# ---------------------------------------------------------------------------
# LLM client factory
# ---------------------------------------------------------------------------

import os
from llm_providers import get_llm_client as _get_llm_client, get_provider_status


@app.get("/api/llm-status")
async def llm_status():
    """Return the current LLM provider configuration status."""
    return get_provider_status()


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------

@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/api/implementation-summary")
async def implementation_summary(
    framework_id: Optional[int] = None,
    session: AsyncSession = Depends(get_session),
):
    """Return counts of implementation statuses."""
    stmt = select(
        Mapping.implementation_status,
        func.count(Mapping.id),
    )
    if framework_id:
        stmt = stmt.join(Control, Mapping.source_control_id == Control.id).where(
            Control.framework_id == framework_id
        )
    stmt = stmt.group_by(Mapping.implementation_status)
    rows = (await session.execute(stmt)).all()

    summary = {"not_assessed": 0, "implemented": 0, "partial": 0, "not_implemented": 0}
    total = 0
    for status, count in rows:
        key = status or "not_assessed"
        summary[key] = count
        total += count

    return {
        "total": total,
        "summary": summary,
        "percentage_implemented": round((summary["implemented"] / total * 100) if total else 0, 1),
    }


# ---------------------------------------------------------------------------
# Frameworks
# ---------------------------------------------------------------------------


@app.get("/api/frameworks", response_model=list[FrameworkOut])
async def list_frameworks(session: AsyncSession = Depends(get_session)):
    """List all frameworks with their control counts."""
    stmt = (
        select(
            Framework,
            func.count(Control.id).label("control_count"),
        )
        .outerjoin(Control, Framework.id == Control.framework_id)
        .group_by(Framework.id)
    )
    rows = (await session.execute(stmt)).all()
    return [
        FrameworkOut(
            id=fw.id,
            name=fw.name,
            short_name=fw.short_name,
            version=fw.version,
            description=fw.description or "",
            is_active=fw.is_active,
            control_count=cnt,
        )
        for fw, cnt in rows
    ]


class FrameworkCreate(BaseModel):
    name: str
    short_name: str
    version: str = ""
    description: str = ""


@app.post("/api/frameworks", response_model=FrameworkOut)
async def create_framework(
    body: FrameworkCreate,
    session: AsyncSession = Depends(get_session),
):
    """Create a new empty framework. Controls can be imported afterwards via /api/import."""
    existing = (await session.execute(
        select(Framework).where(Framework.short_name == body.short_name)
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(400, f"Framework '{body.short_name}' already exists (id={existing.id}).")

    fw = Framework(
        name=body.name,
        short_name=body.short_name,
        version=body.version,
        description=body.description,
        is_active=True,
    )
    session.add(fw)
    await session.commit()
    await session.refresh(fw)
    return FrameworkOut(
        id=fw.id,
        name=fw.name,
        short_name=fw.short_name,
        version=fw.version,
        description=fw.description or "",
        is_active=fw.is_active,
        control_count=0,
    )


# ---------------------------------------------------------------------------
# Controls
# ---------------------------------------------------------------------------


@app.get("/api/controls", response_model=ControlSearchOut)
async def search_controls(
    q: str = "",
    framework_id: Optional[int] = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    session: AsyncSession = Depends(get_session),
):
    """Paginated control search.

    Returns ``{ total, limit, offset, items }`` where ``total`` is the full
    match count (independent of limit/offset) so the UI can render proper
    pagination controls.
    """
    base = select(Control).join(Framework, Control.framework_id == Framework.id)
    if framework_id:
        base = base.where(Control.framework_id == framework_id)
    if q:
        pattern = f"%{q}%"
        base = base.where(
            or_(
                Control.control_id.ilike(pattern),
                Control.title.ilike(pattern),
                Control.description.ilike(pattern),
            )
        )

    count_stmt = select(func.count()).select_from(base.subquery())
    total = (await session.execute(count_stmt)).scalar_one()

    stmt = select(Control, Framework.short_name).join(
        Framework, Control.framework_id == Framework.id
    )
    if framework_id:
        stmt = stmt.where(Control.framework_id == framework_id)
    if q:
        pattern = f"%{q}%"
        stmt = stmt.where(
            or_(
                Control.control_id.ilike(pattern),
                Control.title.ilike(pattern),
                Control.description.ilike(pattern),
            )
        )
        exact_first = case(
            (func.lower(Control.control_id) == q.lower(), 0),
            else_=1,
        )
        stmt = stmt.order_by(exact_first, Control.control_id)
    else:
        stmt = stmt.order_by(Control.control_id)
    stmt = stmt.limit(limit).offset(offset)

    rows = (await session.execute(stmt)).all()
    items = [
        ControlOut(
            id=c.id,
            framework_id=c.framework_id,
            control_id=c.control_id,
            title=c.title or "",
            description=c.description or "",
            category=c.category or "",
            framework_short_name=sn,
        )
        for c, sn in rows
    ]
    return ControlSearchOut(total=total, limit=limit, offset=offset, items=items)


# ---------------------------------------------------------------------------
# Mappings — lookup by control
# ---------------------------------------------------------------------------


@app.get("/api/mappings/{control_id}", response_model=MappingDetail)
async def get_mappings(
    control_id: str,
    framework_id: Optional[int] = None,
    session: AsyncSession = Depends(get_session),
):
    """Return a control and all its mappings. Optionally filter by framework."""
    stmt = select(Control).where(Control.control_id == control_id)
    if framework_id:
        stmt = stmt.where(Control.framework_id == framework_id)
    source = (await session.execute(stmt)).scalar_one_or_none()
    if not source:
        raise HTTPException(404, "Control not found")

    fw = (await session.execute(
        select(Framework.short_name).where(Framework.id == source.framework_id)
    )).scalar_one()

    source_out = ControlOut(
        id=source.id,
        framework_id=source.framework_id,
        control_id=source.control_id,
        title=source.title or "",
        description=source.description or "",
        category=source.category or "",
        framework_short_name=fw,
    )

    # Bidirectional: find mappings where this control is source OR target
    stmt = (
        select(
            Control,
            Framework.short_name,
            Mapping.id,
            Mapping.confidence,
            Mapping.source_type,
            Mapping.source_document,
            Mapping.notes,
            Mapping.implementation_status,
        )
        .join(
            Mapping,
            or_(
                and_(Mapping.source_control_id == source.id, Mapping.target_control_id == Control.id),
                and_(Mapping.target_control_id == source.id, Mapping.source_control_id == Control.id),
            ),
        )
        .join(Framework, Control.framework_id == Framework.id)
        .where(Control.id != source.id)
    )
    rows = (await session.execute(stmt)).all()

    mappings_out = [
        MappingOut(
            id=mid,
            control_id=c.control_id,
            title=c.title or "",
            description=c.description or "",
            category=c.category or "",
            framework_short_name=sn,
            framework_id=c.framework_id,
            confidence=conf,
            source_type=st,
            source_document=sd or "",
            notes=notes or "",
            implementation_status=impl_status or "not_assessed",
        )
        for c, sn, mid, conf, st, sd, notes, impl_status in rows
    ]

    return MappingDetail(source=source_out, mappings=mappings_out)


# ---------------------------------------------------------------------------
# Coverage Analysis
# ---------------------------------------------------------------------------


@app.get("/api/coverage", response_model=CoverageOut)
async def coverage_analysis(
    source: int = Query(..., description="Source framework ID"),
    target: int = Query(..., description="Target framework ID"),
    session: AsyncSession = Depends(get_session),
):
    """Coverage statistics between two frameworks: mapped count, gap count, percentage."""
    src_fw = (await session.execute(
        select(Framework).where(Framework.id == source)
    )).scalar_one_or_none()
    tgt_fw = (await session.execute(
        select(Framework).where(Framework.id == target)
    )).scalar_one_or_none()
    if not src_fw or not tgt_fw:
        raise HTTPException(404, "Framework not found")

    src_controls = (await session.execute(
        select(Control.id, Control.control_id, Control.title).where(Control.framework_id == source)
    )).all()
    tgt_controls = (await session.execute(
        select(Control.id, Control.control_id, Control.title).where(Control.framework_id == target)
    )).all()

    src_ids = {r[0] for r in src_controls}
    src_map = {r[0]: {"id": r[1], "title": r[2] or ""} for r in src_controls}
    tgt_ids = {r[0] for r in tgt_controls}
    tgt_map = {r[0]: {"id": r[1], "title": r[2] or ""} for r in tgt_controls}

    mapped_src = set()
    mapped_tgt = set()

    stmt = select(Mapping.source_control_id, Mapping.target_control_id).where(
        or_(
            and_(Mapping.source_control_id.in_(src_ids), Mapping.target_control_id.in_(tgt_ids)),
            and_(Mapping.target_control_id.in_(src_ids), Mapping.source_control_id.in_(tgt_ids)),
        )
    )
    pairs = (await session.execute(stmt)).all()
    for s, t in pairs:
        if s in src_ids:
            mapped_src.add(s)
            mapped_tgt.add(t)
        else:
            mapped_src.add(t)
            mapped_tgt.add(s)

    total = len(src_controls)
    mapped = len(mapped_src)

    return CoverageOut(
        source_framework=src_fw.short_name,
        target_framework=tgt_fw.short_name,
        total_source_controls=total,
        mapped_controls=mapped,
        unmapped_controls=total - mapped,
        coverage_percentage=round((mapped / total * 100) if total else 0, 1),
        unmapped_control_ids=[src_map[i] for i in src_ids if i not in mapped_src],
        gap_controls=[tgt_map[i] for i in tgt_ids if i not in mapped_tgt],
    )


@app.get("/api/versions/{framework_short_name}/transitions")
async def version_transitions(
    framework_short_name: str,
    session: AsyncSession = Depends(get_session),
):
    """List available version transitions for a framework."""
    fw = (await session.execute(
        select(Framework).where(Framework.short_name == framework_short_name)
    )).scalar_one_or_none()
    if not fw:
        raise HTTPException(404, "Framework not found")

    stmt = (
        select(
            VersionChange.old_version,
            VersionChange.new_version,
            func.count(VersionChange.id).label("change_count"),
        )
        .where(VersionChange.framework_id == fw.id)
        .group_by(VersionChange.old_version, VersionChange.new_version)
        .order_by(VersionChange.old_version)
    )
    rows = (await session.execute(stmt)).all()
    return [
        {"old_version": old, "new_version": new, "change_count": cnt}
        for old, new, cnt in rows
    ]


@app.get("/api/versions/{framework_short_name}/changes", response_model=list[VersionChangeOut])
async def version_changes(
    framework_short_name: str,
    old: str = Query("", alias="from"),
    new: str = Query("", alias="to"),
    session: AsyncSession = Depends(get_session),
):
    fw = (await session.execute(
        select(Framework).where(Framework.short_name == framework_short_name)
    )).scalar_one_or_none()
    if not fw:
        raise HTTPException(404, "Framework not found")

    stmt = select(VersionChange).where(VersionChange.framework_id == fw.id)
    if old:
        stmt = stmt.where(VersionChange.old_version == old)
    if new:
        stmt = stmt.where(VersionChange.new_version == new)
    stmt = stmt.order_by(VersionChange.change_type, VersionChange.old_control_id)

    rows = (await session.execute(stmt)).scalars().all()
    return [
        VersionChangeOut(
            id=r.id,
            old_version=r.old_version,
            new_version=r.new_version,
            change_type=r.change_type,
            old_control_id=r.old_control_id or "",
            new_control_id=r.new_control_id or "",
            description=r.description or "",
            category=r.category or "",
        )
        for r in rows
    ]


class VersionChangeCreate(BaseModel):
    old_version: str
    new_version: str
    change_type: str
    old_control_id: str = ""
    new_control_id: str = ""
    description: str = ""
    category: str = ""


@app.post("/api/versions/{framework_short_name}/changes")
async def add_version_changes(
    framework_short_name: str,
    changes: list[VersionChangeCreate],
    session: AsyncSession = Depends(get_session),
):
    """Bulk-add version change records."""
    fw = (await session.execute(
        select(Framework).where(Framework.short_name == framework_short_name)
    )).scalar_one_or_none()
    if not fw:
        raise HTTPException(404, "Framework not found")

    added = 0
    for ch in changes:
        session.add(VersionChange(
            framework_id=fw.id,
            old_version=ch.old_version,
            new_version=ch.new_version,
            change_type=ch.change_type,
            old_control_id=ch.old_control_id,
            new_control_id=ch.new_control_id,
            description=ch.description,
            category=ch.category,
        ))
        added += 1

    await session.commit()
    return {"added": added}


@app.get("/api/coverage/table")
async def coverage_table(
    source: int = Query(..., description="Source framework ID"),
    target: int = Query(..., description="Target framework ID"),
    session: AsyncSession = Depends(get_session),
):
    """Full mapping table between two frameworks (for export/display)."""
    src_fw = (await session.execute(
        select(Framework).where(Framework.id == source)
    )).scalar_one_or_none()
    tgt_fw = (await session.execute(
        select(Framework).where(Framework.id == target)
    )).scalar_one_or_none()
    if not src_fw or not tgt_fw:
        raise HTTPException(404, "Framework not found")

    src_controls = (await session.execute(
        select(Control).where(Control.framework_id == source).order_by(Control.control_id)
    )).scalars().all()

    rows = []
    for sc in src_controls:
        mapped = (await session.execute(
            select(Control, Mapping.id, Mapping.confidence, Mapping.source_type, Mapping.notes, Mapping.implementation_status)
            .join(
                Mapping,
                or_(
                    and_(Mapping.source_control_id == sc.id, Mapping.target_control_id == Control.id),
                    and_(Mapping.target_control_id == sc.id, Mapping.source_control_id == Control.id),
                ),
            )
            .where(Control.framework_id == target)
        )).all()

        if mapped:
            for tc, mid, conf, st, notes, impl_status in mapped:
                rows.append({
                    "mapping_id": mid,
                    "source_id": sc.control_id,
                    "source_title": sc.title or "",
                    "target_id": tc.control_id,
                    "target_title": tc.title or "",
                    "confidence": conf,
                    "source_type": st,
                    "notes": notes or "",
                    "implementation_status": impl_status or "not_assessed",
                })
        else:
            rows.append({
                "mapping_id": None,
                "source_id": sc.control_id,
                "source_title": sc.title or "",
                "target_id": "",
                "target_title": "",
                "confidence": 0,
                "source_type": "gap",
                "notes": "",
                "implementation_status": "not_assessed",
            })

    return {
        "source_framework": src_fw.short_name,
        "target_framework": tgt_fw.short_name,
        "rows": rows,
    }


@app.get("/api/coverage/export")
async def coverage_export(
    source: int = Query(..., description="Source framework ID"),
    target: int = Query(..., description="Target framework ID"),
    session: AsyncSession = Depends(get_session),
):
    """Download an Excel workbook with mapping summary, full table, and gap list."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    src_fw = (await session.execute(
        select(Framework).where(Framework.id == source)
    )).scalar_one_or_none()
    tgt_fw = (await session.execute(
        select(Framework).where(Framework.id == target)
    )).scalar_one_or_none()
    if not src_fw or not tgt_fw:
        raise HTTPException(404, "Framework not found")

    src_controls = (await session.execute(
        select(Control).where(Control.framework_id == source).order_by(Control.control_id)
    )).scalars().all()
    tgt_controls = (await session.execute(
        select(Control.id, Control.control_id, Control.title).where(Control.framework_id == target)
    )).all()

    tgt_ids = {r[0] for r in tgt_controls}
    tgt_map = {r[0]: {"id": r[1], "title": r[2] or ""} for r in tgt_controls}

    def _confidence_band(value: float) -> str:
        if value >= 0.8:
            return "Strong"
        if value >= 0.5:
            return "Partial"
        if value > 0:
            return "Weak"
        return ""

    table_rows = []
    mapped_src_ids = set()
    mapped_tgt_ids = set()

    for sc in src_controls:
        mapped = (await session.execute(
            select(Control, Mapping.confidence, Mapping.source_type, Mapping.notes)
            .join(
                Mapping,
                or_(
                    and_(Mapping.source_control_id == sc.id, Mapping.target_control_id == Control.id),
                    and_(Mapping.target_control_id == sc.id, Mapping.source_control_id == Control.id),
                ),
            )
            .where(Control.framework_id == target)
        )).all()

        if mapped:
            mapped_src_ids.add(sc.id)
            for tc, conf, st, notes in mapped:
                mapped_tgt_ids.add(tc.id)
                table_rows.append((
                    sc.control_id, sc.title or "",
                    tc.control_id, tc.title or "",
                    st, conf, _confidence_band(conf or 0), notes or "",
                ))
        else:
            table_rows.append((sc.control_id, sc.title or "", "", "", "gap", 0, "", ""))

    total = len(src_controls)
    mapped_count = len(mapped_src_ids)
    pct = round((mapped_count / total * 100) if total else 0, 1)
    unmapped = [(sc.control_id, sc.title or "") for sc in src_controls if sc.id not in mapped_src_ids]
    gap_targets = [(tgt_map[tid]["id"], tgt_map[tid]["title"]) for tid in tgt_ids if tid not in mapped_tgt_ids]

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="C6C6C6"),
    )

    def _style_header(ws, cols):
        for col_idx, val in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 40
    summary_data = [
        ("Source Framework", f"{src_fw.short_name} ({src_fw.version})"),
        ("Target Framework", f"{tgt_fw.short_name} ({tgt_fw.version})"),
        ("Total Source Controls", total),
        ("Mapped Controls", mapped_count),
        ("Unmapped Controls", total - mapped_count),
        ("Coverage", f"{pct}%"),
    ]
    for r, (label, value) in enumerate(summary_data, 1):
        ws_sum.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_sum.cell(row=r, column=2, value=value)

    # Sheet 2: Full Mapping Table
    ws_table = wb.create_sheet("Mapping Table")
    headers = [
        src_fw.short_name, "Source Title",
        tgt_fw.short_name, "Target Title",
        "Type", "Confidence", "Strength", "Notes",
    ]
    _style_header(ws_table, headers)
    gap_fill = PatternFill(start_color="FFF1F1", end_color="FFF1F1", fill_type="solid")
    for r, (s_id, s_title, t_id, t_title, stype, conf, band, notes) in enumerate(table_rows, 2):
        ws_table.cell(row=r, column=1, value=s_id)
        ws_table.cell(row=r, column=2, value=s_title)
        ws_table.cell(row=r, column=3, value=t_id or "No mapping")
        ws_table.cell(row=r, column=4, value=t_title)
        ws_table.cell(row=r, column=5, value=stype)
        ws_table.cell(row=r, column=6, value=round(conf or 0, 2))
        ws_table.cell(row=r, column=7, value=band)
        ws_table.cell(row=r, column=8, value=notes)
        if stype == "gap":
            for col in range(1, 9):
                ws_table.cell(row=r, column=col).fill = gap_fill
    widths = [18, 32, 18, 32, 14, 12, 12, 40]
    for col_idx, w in enumerate(widths, 1):
        ws_table.column_dimensions[chr(64 + col_idx)].width = w

    # Sheet 3: Unmapped Source Controls
    ws_unmapped = wb.create_sheet("Unmapped Controls")
    _style_header(ws_unmapped, ["Control ID", "Title"])
    for r, (ctrl_id, title) in enumerate(unmapped, 2):
        ws_unmapped.cell(row=r, column=1, value=ctrl_id)
        ws_unmapped.cell(row=r, column=2, value=title)
    ws_unmapped.column_dimensions["A"].width = 20
    ws_unmapped.column_dimensions["B"].width = 60

    # Sheet 4: Target Gaps
    ws_gaps = wb.create_sheet("Target Gaps")
    _style_header(ws_gaps, ["Control ID", "Title"])
    for r, (ctrl_id, title) in enumerate(gap_targets, 2):
        ws_gaps.cell(row=r, column=1, value=ctrl_id)
        ws_gaps.cell(row=r, column=2, value=title)
    ws_gaps.column_dimensions["A"].width = 20
    ws_gaps.column_dimensions["B"].width = 60

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"coverage_{src_fw.short_name}_to_{tgt_fw.short_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Document Import
# ---------------------------------------------------------------------------


@app.post("/api/upload", response_model=ParseResult)
async def upload_document(
    file: UploadFile = File(...),
    doc_type: str = Form("BSI Zuordnungstabelle"),
):
    """Parse an uploaded document and return extracted controls/mappings for review."""
    content = await file.read()
    filename = file.filename or ""
    result = parse_uploaded_bytes(content, filename, doc_type)
    return ParseResult(**result)


@app.get("/api/parsers")
async def available_parsers():
    """List registered document parsers (so the UI can populate format dropdowns)."""
    return list_parsers()


@app.post("/api/import", response_model=ImportResult)
async def import_data(
    body: ImportRequest,
    session: AsyncSession = Depends(get_session),
):
    """Persist parsed controls and mappings from a document upload into the database."""
    controls_added = 0
    mappings_added = 0

    src_fw_id = body.source_framework_id
    tgt_fw_id = body.target_framework_id
    source_doc = body.source_document or body.doc_type

    if src_fw_id and tgt_fw_id:
        src_fw = (await session.execute(
            select(Framework).where(Framework.id == src_fw_id)
        )).scalar_one_or_none()
        tgt_fw = (await session.execute(
            select(Framework).where(Framework.id == tgt_fw_id)
        )).scalar_one_or_none()
    else:
        doc_type = body.doc_type
        if "BSI" in doc_type and "C5" not in doc_type:
            short = "BSI"
        elif "C5" in doc_type:
            short = "C5"
        else:
            short = "ISO27001"
        tgt_fw = (await session.execute(
            select(Framework).where(Framework.short_name == short)
        )).scalar_one_or_none()
        src_fw = (await session.execute(
            select(Framework).where(Framework.short_name == "ISO27001")
        )).scalar_one_or_none()

    if not src_fw or not tgt_fw:
        return ImportResult(success=False, error="Source or target framework not found.")

    for ctrl in body.controls:
        existing = (await session.execute(
            select(Control).where(
                Control.framework_id == tgt_fw.id,
                Control.control_id == ctrl["control_id"],
            )
        )).scalar_one_or_none()
        if not existing:
            session.add(Control(
                framework_id=tgt_fw.id,
                control_id=ctrl["control_id"],
                title=ctrl.get("title", ""),
                category=ctrl.get("category", ""),
            ))
            controls_added += 1
    await session.flush()

    all_controls = (await session.execute(
        select(Control.id, Control.control_id, Control.framework_id)
    )).all()
    lookup = {(r[1], r[2]): r[0] for r in all_controls}

    for m in body.mappings:
        s_id = lookup.get((m["source"], src_fw.id))
        t_id = lookup.get((m["target"], tgt_fw.id))

        if not s_id and m["source"]:
            existing = (await session.execute(
                select(Control).where(
                    Control.framework_id == src_fw.id,
                    Control.control_id == m["source"],
                )
            )).scalar_one_or_none()
            if not existing:
                ctrl = Control(
                    framework_id=src_fw.id,
                    control_id=m["source"],
                    title=m["source"],
                    category="",
                )
                session.add(ctrl)
                await session.flush()
                s_id = ctrl.id
                lookup[(m["source"], src_fw.id)] = s_id
            else:
                s_id = existing.id

        if s_id and t_id:
            existing = (await session.execute(
                select(Mapping).where(
                    Mapping.source_control_id == s_id,
                    Mapping.target_control_id == t_id,
                )
            )).scalar_one_or_none()
            if not existing:
                session.add(Mapping(
                    source_control_id=s_id,
                    target_control_id=t_id,
                    confidence=1.0,
                    source_type="official",
                    source_document=source_doc,
                ))
                mappings_added += 1

    await session.commit()
    return ImportResult(success=True, controls_added=controls_added, mappings_added=mappings_added)


# ---------------------------------------------------------------------------
# Mapping CRUD (manual create / edit / delete)
# ---------------------------------------------------------------------------

class MappingDetailOut(BaseModel):
    id: int
    source_control_id: int
    target_control_id: int
    confidence: float
    source_type: str
    source_document: str
    notes: str
    implementation_status: str = "not_assessed"
    owner: str = ""
    review_date: str = ""
    evidence_notes: str = ""


def _serialize_mapping(m: Mapping) -> MappingDetailOut:
    return MappingDetailOut(
        id=m.id,
        source_control_id=m.source_control_id,
        target_control_id=m.target_control_id,
        confidence=m.confidence,
        source_type=m.source_type or "manual",
        source_document=m.source_document or "",
        notes=m.notes or "",
        implementation_status=m.implementation_status or "not_assessed",
        owner=m.owner or "",
        review_date=m.review_date or "",
        evidence_notes=m.evidence_notes or "",
    )


@app.post("/api/mappings", response_model=MappingDetailOut, status_code=201)
async def create_mapping(
    body: MappingCreate,
    session: AsyncSession = Depends(get_session),
):
    """Create a new mapping between two existing controls."""
    if body.source_control_id == body.target_control_id:
        raise HTTPException(400, "Source and target control must differ.")

    src = (await session.execute(
        select(Control).where(Control.id == body.source_control_id)
    )).scalar_one_or_none()
    tgt = (await session.execute(
        select(Control).where(Control.id == body.target_control_id)
    )).scalar_one_or_none()
    if not src or not tgt:
        raise HTTPException(404, "Source or target control not found.")

    # Reject duplicates in either direction.
    existing = (await session.execute(
        select(Mapping).where(
            or_(
                and_(
                    Mapping.source_control_id == body.source_control_id,
                    Mapping.target_control_id == body.target_control_id,
                ),
                and_(
                    Mapping.source_control_id == body.target_control_id,
                    Mapping.target_control_id == body.source_control_id,
                ),
            )
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(409, "A mapping between these controls already exists.")

    confidence = max(0.0, min(1.0, float(body.confidence)))
    mapping = Mapping(
        source_control_id=body.source_control_id,
        target_control_id=body.target_control_id,
        confidence=confidence,
        source_type=body.source_type or "manual",
        source_document=body.source_document or "manual entry",
        notes=body.notes or "",
    )
    session.add(mapping)
    await session.commit()
    await session.refresh(mapping)
    return _serialize_mapping(mapping)


@app.patch("/api/mappings/{mapping_id}", response_model=MappingDetailOut)
async def update_mapping(
    mapping_id: int,
    body: MappingUpdate,
    session: AsyncSession = Depends(get_session),
):
    """Partially update a mapping (confidence, notes, source_type, source_document)."""
    mapping = (await session.execute(
        select(Mapping).where(Mapping.id == mapping_id)
    )).scalar_one_or_none()
    if not mapping:
        raise HTTPException(404, "Mapping not found.")

    if body.confidence is not None:
        mapping.confidence = max(0.0, min(1.0, float(body.confidence)))
    if body.source_type is not None:
        mapping.source_type = body.source_type
    if body.notes is not None:
        mapping.notes = body.notes
    if body.source_document is not None:
        mapping.source_document = body.source_document
    if body.implementation_status is not None:
        mapping.implementation_status = body.implementation_status
    if body.owner is not None:
        mapping.owner = body.owner
    if body.review_date is not None:
        mapping.review_date = body.review_date
    if body.evidence_notes is not None:
        mapping.evidence_notes = body.evidence_notes

    await session.commit()
    await session.refresh(mapping)
    return _serialize_mapping(mapping)


@app.delete("/api/mappings/{mapping_id}", status_code=204)
async def delete_mapping(
    mapping_id: int,
    session: AsyncSession = Depends(get_session),
):
    """Delete a mapping by ID."""
    mapping = (await session.execute(
        select(Mapping).where(Mapping.id == mapping_id)
    )).scalar_one_or_none()
    if not mapping:
        raise HTTPException(404, "Mapping not found.")

    await session.delete(mapping)
    await session.commit()
    return None


# ---------------------------------------------------------------------------
# Compliance Checking API (ARC + RAG frameworks)
# ---------------------------------------------------------------------------

class RegulationUpload(BaseModel):
    name: str
    short_name: str
    version: str = ""
    jurisdiction: str = ""
    full_text: str = ""
    language: str = "en"


class RegulationOut(BaseModel):
    id: int
    name: str
    short_name: str
    version: str
    jurisdiction: str
    language: str


class ComplianceCheckRequest(BaseModel):
    regulation_id: int
    business_text: str


class EventicGraphRequest(BaseModel):
    regulation_id: int


@app.post("/api/regulations/upload", response_model=RegulationOut)
async def upload_regulation(
    body: RegulationUpload,
    session: AsyncSession = Depends(get_session),
):
    """Upload a regulation document for analysis.

    Also creates a Framework and Controls from the extracted statements,
    so they integrate with the existing mapping infrastructure.
    """
    from database import RegulationDocument, Framework, Control
    from arc_pipeline import process_regulation, _split_into_statements

    doc = RegulationDocument(
        name=body.name,
        short_name=body.short_name,
        version=body.version,
        jurisdiction=body.jurisdiction,
        full_text=body.full_text,
        language=body.language,
    )
    session.add(doc)
    await session.flush()

    # Create or get Framework for this regulation
    existing_fw = (await session.execute(
        select(Framework).where(Framework.short_name == body.short_name)
    )).scalar_one_or_none()

    if not existing_fw:
        fw = Framework(
            name=body.name,
            short_name=body.short_name,
            version=body.version or "1.0",
            description=f"Auto-imported regulation: {body.name}",
            is_active=True,
        )
        session.add(fw)
        await session.flush()
        fw_id = fw.id
    else:
        fw_id = existing_fw.id

    # Extract statements and create Controls
    statements = _split_into_statements(body.full_text)
    controls_created = 0
    for idx, stmt in enumerate(statements, 1):
        stmt = stmt.strip()
        if len(stmt) < 15:
            continue
        control_id = f"{body.short_name}-S{idx:03d}"

        existing_ctrl = (await session.execute(
            select(Control).where(
                Control.framework_id == fw_id,
                Control.control_id == control_id,
            )
        )).scalar_one_or_none()

        if not existing_ctrl:
            # Use first 100 chars as title, full statement as description
            title = stmt[:100] + ("..." if len(stmt) > 100 else "")
            session.add(Control(
                framework_id=fw_id,
                control_id=control_id,
                title=title,
                description=stmt,
                category="regulation",
            ))
            controls_created += 1

    await session.commit()
    await session.refresh(doc)
    return RegulationOut(
        id=doc.id, name=doc.name, short_name=doc.short_name,
        version=doc.version or "", jurisdiction=doc.jurisdiction or "",
        language=doc.language or "en",
    )


@app.get("/api/regulations", response_model=list[RegulationOut])
async def list_regulations(session: AsyncSession = Depends(get_session)):
    """List all uploaded regulation documents."""
    from database import RegulationDocument
    stmt = select(RegulationDocument).order_by(RegulationDocument.id)
    rows = (await session.execute(stmt)).scalars().all()
    return [
        RegulationOut(
            id=r.id, name=r.name, short_name=r.short_name,
            version=r.version or "", jurisdiction=r.jurisdiction or "",
            language=r.language or "en",
        )
        for r in rows
    ]


@app.post("/api/regulations/{reg_id}/extract-tuples")
async def extract_regulation_tuples(
    reg_id: int,
    session: AsyncSession = Depends(get_session),
):
    """Run ARC pipeline to extract tuples from a regulation."""
    from database import RegulationDocument, ArcTuple
    from arc_pipeline import process_regulation

    doc = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == reg_id)
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(404, "Regulation not found")

    tuples = process_regulation(doc.full_text)

    db_tuples = []
    for t in tuples:
        arc = ArcTuple(
            regulation_id=reg_id,
            tuple_type=t["tuple_type"],
            source_statement=t.get("source_statement", ""),
            verb=t.get("verb", ""),
            deontic_modal=t.get("deontic_modal", ""),
            sender_phrase=t.get("sender_phrase", ""),
            sender_clause=t.get("sender_clause", ""),
            receiver_phrase=t.get("receiver_phrase", ""),
            receiver_clause=t.get("receiver_clause", ""),
            data_phrase=t.get("data_phrase", ""),
            data_clause=t.get("data_clause", ""),
            transmission_principle=t.get("transmission_principle", ""),
            definiendum=t.get("definiendum", ""),
            definiens=t.get("definiens", ""),
            right_entity=t.get("right_entity", ""),
            right_statement=t.get("right_statement", ""),
        )
        session.add(arc)
        db_tuples.append(t)

    await session.commit()

    return {"regulation_id": reg_id, "tuples": db_tuples, "count": len(db_tuples)}


@app.get("/api/regulations/{reg_id}/tuples")
async def get_regulation_tuples(
    reg_id: int,
    session: AsyncSession = Depends(get_session),
):
    """List extracted ARC tuples for a regulation."""
    from database import RegulationDocument, ArcTuple

    doc = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == reg_id)
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(404, "Regulation not found")

    stmt = select(ArcTuple).where(ArcTuple.regulation_id == reg_id)
    rows = (await session.execute(stmt)).scalars().all()

    tuples = []
    for r in rows:
        tuples.append({
            "id": r.id,
            "tuple_type": r.tuple_type,
            "verb": r.verb,
            "deontic_modal": r.deontic_modal,
            "source_statement": r.source_statement,
        })

    return {"regulation_id": reg_id, "tuples": tuples, "count": len(tuples)}


@app.post("/api/compliance/check")
async def run_compliance_check(
    body: ComplianceCheckRequest,
    session: AsyncSession = Depends(get_session),
):
    """Run RAG-based compliance check against a regulation."""
    from database import RegulationDocument
    from compliance_checker import check_compliance

    doc = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == body.regulation_id)
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(404, "Regulation not found")

    llm_client = _get_llm_client()

    results = check_compliance(
        business_text=body.business_text,
        regulation_text=doc.full_text,
        llm_client=llm_client,
    )

    return {"regulation_id": body.regulation_id, "results": results}


@app.post("/api/eventic-graph/build")
async def build_eventic_graph_endpoint(
    body: EventicGraphRequest,
    session: AsyncSession = Depends(get_session),
):
    """Build eventic graph from a regulation's deontic propositions."""
    from database import RegulationDocument
    from dynamic_layer import extract_deontic_propositions, build_eventic_graph, serialize_graph

    doc = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == body.regulation_id)
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(404, "Regulation not found")

    propositions = extract_deontic_propositions(doc.full_text)
    graph = build_eventic_graph(propositions)
    serialized = serialize_graph(graph)

    return {
        "regulation_id": body.regulation_id,
        "nodes": serialized["nodes"],
        "edges": serialized["edges"],
        "proposition_count": len(propositions),
    }


@app.get("/api/regulations/compare")
async def compare_regulations(
    reg_id_1: int = Query(..., description="First regulation ID"),
    reg_id_2: int = Query(..., description="Second regulation ID"),
    session: AsyncSession = Depends(get_session),
):
    """Compare two regulations using ARCBert phrase similarity."""
    from database import RegulationDocument
    from arc_pipeline import process_regulation, phrase_similarity

    doc1 = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == reg_id_1)
    )).scalar_one_or_none()
    doc2 = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == reg_id_2)
    )).scalar_one_or_none()

    if not doc1 or not doc2:
        raise HTTPException(404, "One or both regulations not found")

    tuples1 = process_regulation(doc1.full_text)
    tuples2 = process_regulation(doc2.full_text)

    if not tuples1 or not tuples2:
        return {
            "reg_id_1": reg_id_1,
            "reg_id_2": reg_id_2,
            "similarity_score": 0.0,
            "matching_pairs": [],
        }

    # Compare statements from both regulations
    matching_pairs = []
    total_sim = 0.0
    pair_count = 0

    for t1 in tuples1[:10]:  # limit for performance
        best_score = 0.0
        best_match = None
        stmt1 = t1.get("source_statement", "")
        if not stmt1:
            continue
        for t2 in tuples2[:10]:
            stmt2 = t2.get("source_statement", "")
            if not stmt2:
                continue
            sim = phrase_similarity(stmt1, stmt2)
            if sim > best_score:
                best_score = sim
                best_match = t2

        if best_match and best_score > 0.3:
            matching_pairs.append({
                "reg1_statement": stmt1,
                "reg2_statement": best_match.get("source_statement", ""),
                "similarity": round(best_score, 4),
            })
            total_sim += best_score
            pair_count += 1

    avg_sim = total_sim / pair_count if pair_count > 0 else 0.0

    return {
        "reg_id_1": reg_id_1,
        "reg_id_2": reg_id_2,
        "similarity_score": round(avg_sim, 4),
        "matching_pairs": matching_pairs,
    }


class GenerateMappingsRequest(BaseModel):
    source_regulation_id: int
    target_regulation_id: int
    threshold: float = 0.4


class FrameworkMappingRequest(BaseModel):
    source_framework_id: int
    target_framework_id: int
    threshold: float = 0.45
    top_k: int = 3


@app.post("/api/frameworks/generate-mappings")
async def generate_framework_mappings(
    body: FrameworkMappingRequest,
    session: AsyncSession = Depends(get_session),
):
    """AI-assisted mapping between two frameworks already in the database.

    Takes every source control, represents it as text (ID + title + description),
    and uses SBERT similarity to find the best-matching target controls above the
    threshold. Results are stored as ai_suggested mappings — identical to what
    the text-based Generate Mappings produces, but works directly from DB controls
    so you don't need to paste thousands of lines of text.
    """
    import asyncio

    src_controls = (await session.execute(
        select(Control).where(Control.framework_id == body.source_framework_id)
    )).scalars().all()

    tgt_controls = (await session.execute(
        select(Control).where(Control.framework_id == body.target_framework_id)
    )).scalars().all()

    if not src_controls or not tgt_controls:
        raise HTTPException(400, "One or both frameworks have no controls in the database.")

    src_fw = (await session.execute(
        select(Framework).where(Framework.id == body.source_framework_id)
    )).scalar_one_or_none()
    tgt_fw = (await session.execute(
        select(Framework).where(Framework.id == body.target_framework_id)
    )).scalar_one_or_none()

    def _build_text(ctrl) -> str:
        parts = [ctrl.control_id, ctrl.title or ""]
        if ctrl.description:
            parts.append(ctrl.description[:300])
        return ": ".join(p for p in parts if p)

    def _run_similarity():
        from arc_pipeline import phrase_similarity
        try:
            from sentence_transformers import SentenceTransformer
            import numpy as np
            model = SentenceTransformer("all-MiniLM-L6-v2")
            src_texts = [_build_text(c) for c in src_controls]
            tgt_texts = [_build_text(c) for c in tgt_controls]
            src_embs = model.encode(src_texts, convert_to_numpy=True, show_progress_bar=False)
            tgt_embs = model.encode(tgt_texts, convert_to_numpy=True, show_progress_bar=False)
            # Cosine similarity matrix
            src_norm = src_embs / (np.linalg.norm(src_embs, axis=1, keepdims=True) + 1e-10)
            tgt_norm = tgt_embs / (np.linalg.norm(tgt_embs, axis=1, keepdims=True) + 1e-10)
            sim_matrix = src_norm @ tgt_norm.T  # (n_src, n_tgt)
            pairs = []
            for i, src_ctrl in enumerate(src_controls):
                row = sim_matrix[i]
                top_indices = np.argsort(row)[::-1][:body.top_k]
                for j in top_indices:
                    score = float(row[j])
                    if score >= body.threshold:
                        pairs.append((src_ctrl, tgt_controls[j], score))
            return pairs
        except Exception:
            # Fallback: phrase_similarity from arc_pipeline
            pairs = []
            for src_ctrl in src_controls:
                src_text = _build_text(src_ctrl)
                scores = []
                for tgt_ctrl in tgt_controls:
                    tgt_text = _build_text(tgt_ctrl)
                    score = phrase_similarity(src_text, tgt_text)
                    if score >= body.threshold:
                        scores.append((tgt_ctrl, score))
                scores.sort(key=lambda x: x[1], reverse=True)
                for tgt_ctrl, score in scores[:body.top_k]:
                    pairs.append((src_ctrl, tgt_ctrl, score))
            return pairs

    pairs = await asyncio.to_thread(_run_similarity)

    # Persist as ai_suggested mappings (skip if already exists)
    existing = set()
    if pairs:
        src_ids = {c.id for c in src_controls}
        tgt_ids = {c.id for c in tgt_controls}
        rows = (await session.execute(
            select(Mapping.source_control_id, Mapping.target_control_id).where(
                Mapping.source_control_id.in_(src_ids),
                Mapping.target_control_id.in_(tgt_ids),
            )
        )).all()
        existing = {(s, t) for s, t in rows}

    added = 0
    for src_ctrl, tgt_ctrl, score in pairs:
        key = (src_ctrl.id, tgt_ctrl.id)
        if key in existing:
            continue
        session.add(Mapping(
            source_control_id=src_ctrl.id,
            target_control_id=tgt_ctrl.id,
            source_type="ai_suggested",
            confidence=round(score, 3),
            source_document=f"AI: {src_fw.short_name if src_fw else '?'} → {tgt_fw.short_name if tgt_fw else '?'}",
            notes=f"Auto-generated by SBERT similarity (score: {score:.2f})",
        ))
        existing.add(key)
        added += 1

    await session.commit()

    src_name = src_fw.short_name if src_fw else str(body.source_framework_id)
    tgt_name = tgt_fw.short_name if tgt_fw else str(body.target_framework_id)

    return {
        "source_framework": src_name,
        "target_framework": tgt_name,
        "source_controls_checked": len(src_controls),
        "target_controls_checked": len(tgt_controls),
        "mappings_added": added,
        "mappings_skipped": len(pairs) - added,
        "threshold": body.threshold,
        "preview": [
            {
                "source": p[0].control_id,
                "source_title": p[0].title or "",
                "target": p[1].control_id,
                "target_title": p[1].title or "",
                "confidence": round(p[2], 3),
            }
            for p in sorted(pairs, key=lambda x: x[2], reverse=True)[:20]
        ],
    }


@app.post("/api/regulations/generate-mappings")
async def generate_regulation_mappings(
    body: GenerateMappingsRequest,
    session: AsyncSession = Depends(get_session),
):
    """Automatically generate mapping suggestions between two regulations.

    Uses SBERT semantic similarity to find matching statements, then
    stores them as ai_suggested mappings in the existing mappings table.
    """
    from database import RegulationDocument, Framework, Control
    from mapping_engine import generate_mappings, format_as_suggestions

    doc1 = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == body.source_regulation_id)
    )).scalar_one_or_none()
    doc2 = (await session.execute(
        select(RegulationDocument).where(RegulationDocument.id == body.target_regulation_id)
    )).scalar_one_or_none()

    if not doc1 or not doc2:
        raise HTTPException(404, "One or both regulations not found")

    mappings = generate_mappings(
        source_text=doc1.full_text,
        target_text=doc2.full_text,
        threshold=body.threshold,
    )

    suggestions = format_as_suggestions(
        mappings,
        source_reg_name=doc1.short_name,
        target_reg_name=doc2.short_name,
    )

    # Persist mappings into the existing mappings table
    # Find frameworks for both regulations
    src_fw = (await session.execute(
        select(Framework).where(Framework.short_name == doc1.short_name)
    )).scalar_one_or_none()
    tgt_fw = (await session.execute(
        select(Framework).where(Framework.short_name == doc2.short_name)
    )).scalar_one_or_none()

    persisted_count = 0
    if src_fw and tgt_fw:
        # Load all controls for both frameworks
        src_controls = (await session.execute(
            select(Control).where(Control.framework_id == src_fw.id)
        )).scalars().all()
        tgt_controls = (await session.execute(
            select(Control).where(Control.framework_id == tgt_fw.id)
        )).scalars().all()

        # Index by description for matching
        src_by_desc = {c.description: c for c in src_controls if c.description}
        tgt_by_desc = {c.description: c for c in tgt_controls if c.description}

        for suggestion in suggestions:
            src_stmt = suggestion["source_statement"]
            tgt_stmt = suggestion["target_statement"]

            src_ctrl = src_by_desc.get(src_stmt)
            tgt_ctrl = tgt_by_desc.get(tgt_stmt)

            if src_ctrl and tgt_ctrl:
                # Check if mapping already exists
                existing = (await session.execute(
                    select(Mapping).where(
                        or_(
                            and_(
                                Mapping.source_control_id == src_ctrl.id,
                                Mapping.target_control_id == tgt_ctrl.id,
                            ),
                            and_(
                                Mapping.source_control_id == tgt_ctrl.id,
                                Mapping.target_control_id == src_ctrl.id,
                            ),
                        )
                    )
                )).scalar_one_or_none()

                if not existing:
                    session.add(Mapping(
                        source_control_id=src_ctrl.id,
                        target_control_id=tgt_ctrl.id,
                        confidence=suggestion["confidence"],
                        source_type="ai_suggested",
                        source_document=suggestion["source_document"],
                        notes=suggestion["notes"],
                    ))
                    persisted_count += 1

        await session.commit()

    return {
        "source_regulation": doc1.short_name,
        "target_regulation": doc2.short_name,
        "threshold": body.threshold,
        "mappings_found": len(suggestions),
        "mappings_persisted": persisted_count,
        "mappings": suggestions,
    }


# ---------------------------------------------------------------------------
# Embedding & Suggestion Endpoints (scaffolding — AI module extension point)
# ---------------------------------------------------------------------------

class EmbeddingRequest(BaseModel):
    framework_id: int
    model: str = "text-embedding-3-small"


@app.post("/api/embeddings/generate")
async def generate_embeddings(body: EmbeddingRequest):
    """Placeholder for embedding generation — to be implemented by the AI module."""
    return {
        "status": "not_implemented",
        "message": "Embedding generation will be handled by the AI module. "
                   f"Requested: framework_id={body.framework_id}, model={body.model}",
    }


@app.get("/api/mappings/suggest")
async def suggest_mappings(
    control_id: str = Query(..., description="Source control ID to find suggestions for"),
    framework_id: int = Query(..., description="Target framework ID"),
    top_k: int = Query(5, description="Number of suggestions to return"),
    session: AsyncSession = Depends(get_session),
):
    """AI-powered mapping suggestions via pgvector cosine similarity.

    Returns empty results until embeddings are generated by the AI module.
    """
    source = (await session.execute(
        select(Control).where(Control.control_id == control_id)
    )).scalar_one_or_none()
    if not source:
        raise HTTPException(404, "Source control not found")

    if source.embedding is None:
        return {
            "source_control": control_id,
            "target_framework_id": framework_id,
            "suggestions": [],
            "message": "No embeddings available. Run embedding generation first.",
        }

    try:
        from pgvector.sqlalchemy import Vector
        stmt = (
            select(
                Control.control_id,
                Control.title,
                Control.description,
                Control.embedding.cosine_distance(source.embedding).label("distance"),
            )
            .where(Control.framework_id == framework_id)
            .where(Control.embedding.isnot(None))
            .order_by("distance")
            .limit(top_k)
        )
        rows = (await session.execute(stmt)).all()
        suggestions = [
            {
                "control_id": r[0],
                "title": r[1] or "",
                "description": r[2] or "",
                "confidence": round(1 - r[3], 4),
                "source_type": "ai_suggested",
            }
            for r in rows
        ]
    except Exception:
        suggestions = []

    return {
        "source_control": control_id,
        "target_framework_id": framework_id,
        "suggestions": suggestions,
        "message": "" if suggestions else "No embeddings available. Run embedding generation first.",
    }


# ---------------------------------------------------------------------------
# Policy Gap Check
# ---------------------------------------------------------------------------

class PolicyCheckRequest(BaseModel):
    source_framework_id: int
    target_framework_id: int
    policy_text: str
    threshold: float = 0.45


@app.post("/api/coverage/check-policy")
async def check_policy_against_gaps(
    body: PolicyCheckRequest,
    session: AsyncSession = Depends(get_session),
):
    """Check how well a policy document covers the unmapped controls.

    Uses the full ARC pipeline (Algorithm 1 + eventic graph):
    1. Synthesizes each unmapped control into a deontic "shall" clause.
    2. Runs extract_deontic_propositions + build_eventic_graph over them.
    3. Chunks the policy text and matches each chunk against the graph
       using SBERT + the compliance reasoning layer (LLM if available,
       rule-based otherwise).
    4. Aggregates per-control: covered / possibly_covered / not_covered.
    """
    import asyncio

    src_controls = (await session.execute(
        select(Control).where(Control.framework_id == body.source_framework_id)
    )).scalars().all()

    tgt_control_ids = set((await session.execute(
        select(Control.id).where(Control.framework_id == body.target_framework_id)
    )).scalars().all())

    src_id_set = {c.id for c in src_controls}
    mapped_src: set[int] = set()

    if src_id_set and tgt_control_ids:
        pairs = (await session.execute(
            select(Mapping.source_control_id, Mapping.target_control_id).where(
                or_(
                    and_(
                        Mapping.source_control_id.in_(src_id_set),
                        Mapping.target_control_id.in_(tgt_control_ids),
                    ),
                    and_(
                        Mapping.target_control_id.in_(src_id_set),
                        Mapping.source_control_id.in_(tgt_control_ids),
                    ),
                )
            )
        )).all()
        for s, t in pairs:
            if s in src_id_set:
                mapped_src.add(s)
            else:
                mapped_src.add(t)

    unmapped = [c for c in src_controls if c.id not in mapped_src]

    if not unmapped or not body.policy_text.strip():
        return {
            "unmapped_count": len(unmapped),
            "checked_count": 0,
            "covered_count": 0,
            "possibly_covered_count": 0,
            "not_covered_count": 0,
            "results": [],
            "message": "No unmapped controls or no policy text provided.",
        }

    def _run_matching():
        """Run the ARC pipeline against the policy text.

        Each unmapped control is synthesized into a regulation clause so that
        extract_deontic_propositions can parse it into obligations.  All control
        clauses are combined into one eventic graph; the policy text is chunked
        and each chunk is matched against the graph.  We then aggregate per-
        control: a control is "covered" when at least one policy chunk produced
        a compliant or undetermined judgment for one of its obligations.
        """
        from dynamic_layer import extract_deontic_propositions, build_eventic_graph, chunk_text
        from static_layer import extract_term_definitions
        from compliance_checker import match_chunk_to_graph, fuse_knowledge, _reason_rule_based
        from llm_providers import get_llm_client, get_reasoning_fn
        from compliance_checker import build_compliance_prompt, _reason_with_llm

        # ── Step 1: synthesize each control into a "shall" clause ──────────
        # Maps action_node_text → control so we can trace matches back.
        clause_to_ctrl: dict[str, object] = {}

        def _synthesize(ctrl) -> str:
            title = ctrl.title or ctrl.control_id
            desc = ctrl.description or ""
            # Short description: use it directly as the obligation body
            body_text = desc[:300] if desc else f"implement {title.lower()}"
            clause = f"The organization shall {body_text}" if not any(
                w in body_text.lower() for w in ("shall", "must", "should")
            ) else body_text
            return clause

        all_clauses = ""
        ctrl_clauses: dict[int, str] = {}  # ctrl.id → clause text
        for ctrl in unmapped[:60]:
            clause = _synthesize(ctrl)
            ctrl_clauses[ctrl.id] = clause
            all_clauses += clause + "\n"

        if not all_clauses.strip():
            return []

        # ── Step 2: build eventic graph from synthesized regulation text ───
        propositions = extract_deontic_propositions(all_clauses)
        graph = build_eventic_graph(propositions)
        definitions = extract_term_definitions(all_clauses)

        # ── Step 3: chunk the policy document ─────────────────────────────
        policy_chunks = [c for c in chunk_text(body.policy_text) if len(c.strip()) > 20]
        if not policy_chunks:
            return []

        # ── Step 4: match every policy chunk against the graph ─────────────
        # Collect all matched obligations per chunk so we can trace which
        # control's obligation was satisfied.
        llm_client = None
        try:
            llm_client = get_llm_client()
        except Exception:
            pass

        chunk_results = []
        for chunk in policy_chunks:
            matches = match_chunk_to_graph(chunk, graph, threshold=0.2)
            static_knowledge = [
                d for d in definitions
                if any(d.get("term", "").lower() in chunk.lower().split())
            ]
            knowledge = fuse_knowledge(static_knowledge, matches)

            if llm_client and get_reasoning_fn(llm_client):
                prompt = build_compliance_prompt(chunk, knowledge)
                verdict = _reason_with_llm(prompt, llm_client)
            else:
                verdict = _reason_rule_based(chunk, knowledge, matches)

            chunk_results.append({
                "chunk": chunk,
                "judgment": verdict["judgment"],
                "explanation": verdict["explanation"],
                "matched_obligations": [m["text"] for m in matches if m.get("relation") == "duty"],
                "top_match_score": max((m["score"] for m in matches), default=0.0),
            })

        # ── Step 5: aggregate per control ─────────────────────────────────
        results = []
        for ctrl in unmapped[:60]:
            clause = ctrl_clauses[ctrl.id]
            clause_lower = clause.lower()

            # Find policy chunks whose matched obligations overlap with this
            # control's synthesized clause.
            best_judgment = "not_covered"
            best_score = 0.0
            best_chunk = ""
            best_explanation = ""

            for cr in chunk_results:
                for obligation in cr["matched_obligations"]:
                    # Check if this obligation text appears in or is close to
                    # this control's clause (word overlap as a lightweight check)
                    obl_words = set(obligation.lower().split())
                    clause_words = set(clause_lower.split())
                    overlap_ratio = len(obl_words & clause_words) / max(len(obl_words), 1)

                    if overlap_ratio < 0.15:
                        continue

                    score = cr["top_match_score"]
                    judgment = cr["judgment"]

                    # Promote coverage verdict
                    if judgment == "compliant" and best_judgment != "covered":
                        best_judgment = "covered"
                        best_score = score
                        best_chunk = cr["chunk"]
                        best_explanation = cr["explanation"]
                    elif judgment == "undetermined" and best_judgment == "not_covered":
                        best_judgment = "possibly_covered"
                        best_score = score
                        best_chunk = cr["chunk"]
                        best_explanation = cr["explanation"]
                    elif score > best_score:
                        best_score = score
                        best_chunk = cr["chunk"]
                        best_explanation = cr["explanation"]

            results.append({
                "control_id": ctrl.control_id,
                "title": ctrl.title or "",
                "coverage": best_judgment,
                "confidence": round(best_score, 3),
                "matching_snippet": best_chunk[:250] if best_judgment != "not_covered" else "",
                "explanation": best_explanation,
            })

        return sorted(results, key=lambda x: x["confidence"], reverse=True)

    results = await asyncio.to_thread(_run_matching)
    if results is None:
        results = []

    covered = sum(1 for r in results if r["coverage"] == "covered")
    possibly = sum(1 for r in results if r["coverage"] == "possibly_covered")
    not_covered = sum(1 for r in results if r["coverage"] == "not_covered")

    return {
        "unmapped_count": len(unmapped),
        "checked_count": len(results),
        "covered_count": covered,
        "possibly_covered_count": possibly,
        "not_covered_count": not_covered,
        "results": results,
    }


# ---------------------------------------------------------------------------
# Static files – served AFTER API routes so /api/* takes priority
# ---------------------------------------------------------------------------

app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def index():
    return FileResponse("static/index.html")
